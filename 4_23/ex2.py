import openpyxl

df = openpyxl.load_Workbook('resource/excel-data.xlsx')

wb.active
### xの値でループ（xの値を2から最大行番号までループ）
# for x in range(ws.max_row):

#         #abc.xlsxファイルのdefシートから各科目ごとに全員の合計点を計算
#         dict_text[]= dict_text["英語"] + ws.cell(x, 2).value
#         dict_text["math score"]= dict_text["math score"] + ws.cell(x, 3).value
#         dict_text["国語"]= dict_text["国語"] + ws.cell(x, 4).value

# #各科目ごとに平均点を求めて、abc2.xlsxファイルのdef2シートの平均点欄へ書込み
# ws.cell(2, 2).value = dict_text["英語"] / (ws.max_row - 1)
# ws.cell(2, 3).value = dict_text["math score"] / (ws.max_row - 1)
# ws.cell(2, 4).value = dict_text["国語"] / (ws.max_row - 1)